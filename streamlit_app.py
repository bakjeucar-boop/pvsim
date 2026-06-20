import copy
import datetime
import io

import altair as alt
import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from mg_weather_openmeteo import getweatherdatamixed, resolvetimezoneandelevation
from mg_pv_core import (
    LOSSPARAMS,
    GeneratorConfig,
    computepvforgenerator,
    defaultlosssettings,
    defaultobstacles,
    lossparamsforgenerator,
)


st.set_page_config(
    page_title="PV-Forecast",
    layout="wide",
    page_icon="☀️",
    initial_sidebar_state="expanded",
)


FIELD_HELP_TEXTS = {
    "gamma": (
        "태양광 모듈 온도가 기준온도보다 올라갈 때 출력이 얼마나 감소하는지를 나타내는 값입니다.\n\n"
        "추천값:\n"
        "- 다결정 실리콘 모듈: -0.40 ~ -0.45 %/°C\n"
        "- 단결정 실리콘 모듈: -0.30 ~ -0.40 %/°C\n"
        "- 일반 결정질 실리콘 모듈: -0.35 %/°C\n"
        "- 화합물 반도체(CIGS/CdTe): -0.20 ~ -0.35 %/°C\n"
        "- 비정질 실리콘(a-Si): -0.15 ~ -0.25 %/°C"
    ),
    "bifaciality": (
        "양면형 모듈의 후면 발전 효율을 나타내는 값입니다. 70%는 후면으로 들어오는 "
        "반사광을 전면과 같은 효율로 100% 발전하는 것이 아니라, 전면 발전량 대비 약 "
        "70% 수준의 효율로 전기를 생산한다는 의미입니다. 단면 모듈에서는 계산에 반영되지 않습니다.\n\n"
        "추천값:\n"
        "- 일반 양면형 모듈: 65 ~ 75%\n"
        "- 고성능 양면형 모듈: 75 ~ 85%\n"
        "- 보수적인 추정: 60 ~ 70%\n"
        "- 현재 기본값: 70%"
    ),
    "azimuth": (
        "태양광 모듈이 바라보는 방향입니다. 이 코드에서는 0도=북쪽, 90도=동쪽, "
        "180도=남쪽, 270도=서쪽 기준으로 입력합니다.\n\n"
        "추천값:\n"
        "- 남향: 180도\n"
        "- 남동향: 135도\n"
        "- 남서향: 225도\n"
        "- 동향: 90도\n"
        "- 서향: 270도\n"
        "- 북향: 0도"
    ),
    "albedo": (
        "지표면이 햇빛을 얼마나 반사하는지를 나타내는 값입니다. 0에 가까울수록 반사가 적고, "
        "1에 가까울수록 반사가 많습니다. 양면형 모듈에서는 지면 반사광이 후면 발전량에 영향을 줍니다.\n\n"
        "추천값:\n"
        "- 아스팔트: 0.05 ~ 0.12\n"
        "- 흙/일반 지면: 0.15 ~ 0.25\n"
        "- 잔디: 0.20 ~ 0.30\n"
        "- 콘크리트: 0.25 ~ 0.40\n"
        "- 밝은 콘크리트/흰색 지붕: 0.40 ~ 0.60\n"
        "- 눈 덮인 지면: 0.60 ~ 0.90"
    ),
}

LOSS_HELP_TEXTS = {
    "soiling": (
        "모듈 표면에 먼지, 꽃가루, 새 배설물, 오염물 등이 쌓여 발생하는 손실입니다.\n\n"
        "추천값:\n"
        "- 관리가 잘 되는 설비: 1 ~ 2%\n"
        "- 일반적인 설비: 2 ~ 4%\n"
        "- 먼지가 많은 지역/청소가 드문 경우: 4 ~ 8%"
    ),
    "mismatch": (
        "모듈 간 출력 차이 때문에 발생하는 손실입니다. 같은 설비 안에서도 각 모듈의 성능이 조금씩 달라 "
        "전체 출력이 줄어들 수 있습니다.\n\n"
        "추천값:\n"
        "- 모듈 품질/배치가 좋은 경우: 0.5 ~ 1%\n"
        "- 일반적인 설비: 1 ~ 2%"
    ),
    "wiring": (
        "전선에서 전기가 이동하면서 발생하는 손실입니다. 전선 길이가 길거나 전선 굵기가 부족하면 손실이 커집니다.\n\n"
        "추천값:\n"
        "- 설계가 좋은 경우: 1% 내외\n"
        "- 일반적인 설비: 1 ~ 3%\n"
        "- 전선 거리가 긴 경우: 3% 이상"
    ),
    "connections": (
        "커넥터, 접속함, 단자 등 전기 연결부에서 발생하는 손실입니다.\n\n"
        "추천값:\n"
        "- 관리 상태가 좋은 경우: 0.5%\n"
        "- 일반적인 설비: 0.5 ~ 1%\n"
        "- 오래되었거나 접속부가 많은 경우: 1 ~ 2%"
    ),
    "lid": (
        "Light Induced Degradation의 약자로, 모듈이 처음 햇빛에 노출된 뒤 초기 성능이 약간 감소하는 현상입니다.\n\n"
        "추천값:\n"
        "- LID 저감 모듈: 0.5 ~ 1%\n"
        "- 일반 결정질 실리콘 모듈: 1 ~ 2%"
    ),
    "nameplate_rating": (
        "모듈에 표시된 정격출력과 실제 출력 사이의 차이로 인한 손실입니다.\n\n"
        "추천값:\n"
        "- 양의 출력공차 모듈 사용 시: 0%\n"
        "- 일반적인 설비: 0 ~ 1%\n"
        "- 보수적으로 계산할 때: 1%"
    ),
    "age": (
        "태양광 모듈이 시간이 지나면서 성능이 서서히 감소하는 손실입니다. 이 항목은 현재 계산에 바로 적용되는 "
        "누적 손실률로 입력하는 것이 자연스럽습니다.\n\n"
        "추천값:\n"
        "- 신규 설비: 0%\n"
        "- 설치 5년차: 약 2.5%\n"
        "- 설치 10년차: 약 5%\n"
        "- 일반적인 계산식: 사용연수 × 0.5%"
    ),
    "availability": (
        "고장, 점검, 인버터 정지, 계통 문제 등으로 설비가 정상 가동하지 못하는 시간에 따른 손실입니다.\n\n"
        "추천값:\n"
        "- 관리가 잘 되는 상업용 설비: 0.5 ~ 2%\n"
        "- 일반적인 설비: 1 ~ 3%\n"
        "- 정전/고장이 잦은 경우: 3% 이상"
    ),
}


def make_default_generator(name=None):
    return GeneratorConfig(
        name=name or "Generator 1",
        obstacles=[],
        losssettings=defaultlosssettings(),
    )


def clone_generator(gen):
    return GeneratorConfig(
        name=gen.name,
        modulepdcstcw=gen.modulepdcstcw,
        modulecount=gen.modulecount,
        invacratedwper=gen.invacratedwper,
        invertercount=gen.invertercount,
        etainvnom=gen.etainvnom,
        surfacetilt=gen.surfacetilt,
        surfaceazimuth=gen.surfaceazimuth,
        albedo=gen.albedo,
        mounting=gen.mounting,
        facetype=gen.facetype,
        bifacialityfactorpct=gen.bifacialityfactorpct,
        gammapctperc=gen.gammapctperc,
        plannedavailability=gen.plannedavailability,
        obstacles=copy.deepcopy(gen.obstacles or []),
        losssettings=copy.deepcopy(gen.losssettings or defaultlosssettings()),
    )


def load_form_from_generator(gen, edit_idx=None):
    st.session_state.form_version = st.session_state.get("form_version", 0) + 1
    purge_form_widget_state()
    st.session_state.show_loss_dialog = False
    st.session_state.form_edit_idx = edit_idx
    st.session_state.form_gen = clone_generator(gen)
    obstacles = [
        obs
        for obs in (gen.obstacles or [])
        if obs.get("enabled")
        or any(obs.get(k) not in (None, "", 0, 0.0) for k in ("centerazdeg", "distm", "heightm", "widthm"))
    ]
    st.session_state.form_obstacles = copy.deepcopy(obstacles)
    st.session_state.form_losssettings = copy.deepcopy(
        gen.losssettings or defaultlosssettings()
    )


def reset_form():
    st.session_state.form_version = st.session_state.get("form_version", 0) + 1
    purge_form_widget_state()
    st.session_state.show_loss_dialog = False
    st.session_state.form_edit_idx = None
    st.session_state.form_gen = make_default_generator(
        f"Generator {len(st.session_state.generators) + 1}"
    )
    st.session_state.form_obstacles = []
    st.session_state.form_losssettings = defaultlosssettings()


def refresh_form_widgets():
    st.session_state.form_version = st.session_state.get("form_version", 0) + 1


def add_form_obstacle():
    st.session_state.form_obstacles.append(
        {
            "enabled": True,
            "centerazdeg": 0.0,
            "distm": 0.0,
            "heightm": 0.0,
            "widthm": 0.0,
        }
    )
    refresh_form_widgets()


def delete_form_obstacle(idx):
    if 0 <= idx < len(st.session_state.form_obstacles):
        st.session_state.form_obstacles.pop(idx)
    refresh_form_widgets()


def is_old_default_obstacle(obs):
    default_obs = defaultobstacles()[0]
    return all(obs.get(key) == value for key, value in default_obs.items())


def purge_form_widget_state():
    for key in list(st.session_state.keys()):
        key_s = str(key)
        parts = key_s.split("_")
        is_versioned_form_key = (
            len(parts) > 2 and parts[0] == "form" and parts[1].isdigit()
        )
        if is_versioned_form_key:
            del st.session_state[key]


def first_existing_col(df, candidates):
    for col in candidates:
        if col in df.columns:
            return col
    return None


def build_generator_from_form():
    base = st.session_state.form_gen
    return GeneratorConfig(
        name=base.name,
        modulepdcstcw=base.modulepdcstcw,
        modulecount=int(base.modulecount),
        invacratedwper=base.invacratedwper,
        invertercount=int(base.invertercount),
        etainvnom=base.etainvnom,
        surfacetilt=base.surfacetilt,
        surfaceazimuth=base.surfaceazimuth,
        albedo=base.albedo,
        mounting=base.mounting,
        facetype=base.facetype,
        bifacialityfactorpct=base.bifacialityfactorpct,
        gammapctperc=base.gammapctperc,
        plannedavailability=base.plannedavailability,
        obstacles=copy.deepcopy(st.session_state.form_obstacles),
        losssettings=copy.deepcopy(st.session_state.form_losssettings),
    )


def generator_summary_rows():
    rows = []
    for idx, gen in enumerate(st.session_state.generators, start=1):
        rows.append(
            {
                "번호": idx,
                "이름": gen.name,
                "모듈 STC [W]": gen.modulepdcstcw,
                "모듈 수": gen.modulecount,
                "DC 용량 [kW]": round(gen.pdc0totalw() / 1000.0, 2),
                "인버터 용량 [kW]": round(gen.invactotalw() / 1000.0, 2),
                "방위각 [deg]": gen.surfaceazimuth,
                "경사각 [deg]": gen.surfacetilt,
                "모듈 타입": gen.facetype,
                "장애물 수": sum(1 for obs in (gen.obstacles or []) if obs.get("enabled")),
            }
        )
    return rows


def build_generator_detail_summary(results=None):
    results = results or st.session_state.results or {}
    rows = []
    for gen in st.session_state.generators:
        generation_kwh = 0.0
        peak_kwh = 0.0
        if gen.name in results:
            hourly = results[gen.name]["hourly"]
            generation_kwh = float(hourly["generationkwh"].sum())
            peak_kwh = float(hourly["generationkwh"].max())
        rows.append(
            {
                "발전기": gen.name,
                "발전량 [kWh]": round(generation_kwh, 2),
                "DC 용량 [kW]": round(gen.pdc0totalw() / 1000.0, 2),
                "인버터 용량 [kW]": round(gen.invactotalw() / 1000.0, 2),
                "방위각 [deg]": round(float(gen.surfaceazimuth), 2),
                "경사각 [deg]": round(float(gen.surfacetilt), 2),
                "모듈 타입": gen.facetype,
            }
        )
    if rows:
        rows.append(
            {
                "발전기": "합계",
                "발전량 [kWh]": round(sum(row["발전량 [kWh]"] for row in rows), 2),
                "DC 용량 [kW]": round(sum(row["DC 용량 [kW]"] for row in rows), 2),
                "인버터 용량 [kW]": round(
                    sum(row["인버터 용량 [kW]"] for row in rows), 2
                ),
                "방위각 [deg]": "",
                "경사각 [deg]": "",
                "모듈 타입": "",
            }
        )
    return pd.DataFrame(rows)


def weather_export_frame(index):
    weatherhourly = st.session_state.get("weatherhourly")
    df = pd.DataFrame(index=index)
    if weatherhourly is None:
        return df
    columns = [
        ("shortwave_radiation", "전일사량_GHI_Wm2"),
        ("shortwaveradiation", "전일사량_GHI_Wm2"),
        ("direct_normal_irradiance", "직달일사량_DNI_Wm2"),
        ("directnormalirradiance", "직달일사량_DNI_Wm2"),
        ("diffuse_radiation", "산란일사량_DHI_Wm2"),
        ("diffuseradiation", "산란일사량_DHI_Wm2"),
        ("temperature_2m", "외기온도_C"),
        ("temperature2m", "외기온도_C"),
    ]
    used_labels = set()
    for source_col, label in columns:
        if source_col in weatherhourly.columns and label not in used_labels:
            df[label] = weatherhourly[source_col].reindex(index).values
            used_labels.add(label)
    return df


def clean_hourly_export_frame(name, hourly, include_weather=True):
    out = pd.DataFrame(index=hourly.index)
    out["발전기"] = name
    out["발전량_kWh"] = hourly["generationkwh"].values
    out["AC출력_kW"] = hourly["acpowerw"].values / 1000.0
    if "isshaded" in hourly.columns:
        out["음영여부"] = hourly["isshaded"].astype(bool).values
    if include_weather:
        out = pd.concat([out, weather_export_frame(hourly.index)], axis=1)
    try:
        out.index = out.index.tz_localize(None)
    except Exception:
        pass
    out.index.name = "날짜시각"
    return out


def build_csv_export(results):
    frames = [
        clean_hourly_export_frame(name, data["hourly"], include_weather=True)
        for name, data in results.items()
    ]
    csv_df = pd.concat(frames).reset_index()
    return csv_df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")


def style_worksheet(ws, freeze_cell="A2"):
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = freeze_cell
    ws.auto_filter.ref = ws.dimensions
    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        max_len = 10
        for cell in ws[column_letter]:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)), 28))
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
        ws.column_dimensions[column_letter].width = max_len + 2


def style_summary_worksheet(ws, detail_header_row=11):
    title_fill = PatternFill("solid", fgColor="1F4E79")
    header_fill = PatternFill("solid", fgColor="5B9BD5")
    ws["A1"].fill = title_fill
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(2, ws.max_column))
    for row_idx in range(3, 9):
        ws.cell(row_idx, 1).font = Font(bold=True)
        ws.cell(row_idx, 2).number_format = "#,##0.00"
    for cell in ws[detail_header_row]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = f"A{detail_header_row + 1}"
    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        max_len = 10
        for cell in ws[column_letter]:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)), 28))
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
        ws.column_dimensions[column_letter].width = max_len + 2


def write_dataframe(ws, df, start_row=1, start_col=1, include_index=False):
    if include_index:
        df = df.reset_index()
    for col_offset, col_name in enumerate(df.columns):
        ws.cell(start_row, start_col + col_offset, col_name)
    for row_offset, row in enumerate(df.itertuples(index=False), start=1):
        for col_offset, value in enumerate(row):
            ws.cell(start_row + row_offset, start_col + col_offset, value)
    style_worksheet(ws)


def generate_excel_export(results, start_date, end_date, lat, lon):
    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "요약"
    ws_summary["A1"] = "태양광 발전량 예측 결과"
    ws_summary["A1"].font = Font(bold=True, size=14)
    summary_rows = [
        ("위도", lat),
        ("경도", lon),
        ("시작일", start_date.strftime("%Y-%m-%d")),
        ("종료일", end_date.strftime("%Y-%m-%d")),
        ("총 발전량 [kWh]", float(results["Total"]["hourly"]["generationkwh"].sum())),
        ("일평균 발전량 [kWh/day]", float(results["Total"]["daily"]["dailygenerationkwh"].mean())),
    ]
    for row_idx, (label, value) in enumerate(summary_rows, start=3):
        ws_summary.cell(row_idx, 1, label)
        ws_summary.cell(row_idx, 2, value)
    detail_df = build_generator_detail_summary(results)
    for col_idx, col_name in enumerate(detail_df.columns, start=1):
        ws_summary.cell(11, col_idx, col_name)
    for row_offset, row in enumerate(detail_df.itertuples(index=False), start=12):
        for col_idx, value in enumerate(row, start=1):
            ws_summary.cell(row_offset, col_idx, value)
    style_summary_worksheet(ws_summary)

    total_hourly = clean_hourly_export_frame(
        "합계", results["Total"]["hourly"], include_weather=True
    ).drop(columns=["발전기"], errors="ignore")
    ws_total = wb.create_sheet("통합_시간별")
    write_dataframe(ws_total, total_hourly, include_index=True)

    generator_frames = []
    for name, data in results.items():
        if name == "Total":
            continue
        generator_frames.append(clean_hourly_export_frame(name, data["hourly"], include_weather=False))
    if generator_frames:
        generator_hourly = pd.concat(generator_frames).reset_index()
        ws_gen_hourly = wb.create_sheet("발전기별_시간별")
        write_dataframe(ws_gen_hourly, generator_hourly)

    daily_rows = []
    for name, data in results.items():
        daily = data["daily"].copy()
        try:
            daily.index = daily.index.tz_localize(None)
        except Exception:
            pass
        for ts, row in daily.iterrows():
            daily_rows.append(
                {
                    "날짜": ts.date() if hasattr(ts, "date") else ts,
                    "발전기": "합계" if name == "Total" else name,
                    "일 발전량_kWh": float(row["dailygenerationkwh"]),
                }
            )
    ws_daily = wb.create_sheet("일별_요약")
    write_dataframe(ws_daily, pd.DataFrame(daily_rows))

    settings_df = pd.DataFrame(generator_summary_rows())
    ws_settings = wb.create_sheet("발전기_설정")
    write_dataframe(ws_settings, settings_df)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def render_hourly_result_chart(hourly, weatherhourly, ghi_col, temp_col):
    chart_df = pd.DataFrame(
        {
            "time": hourly.index,
            "generationkwh": hourly["generationkwh"].values,
        }
    )
    if weatherhourly is not None and ghi_col:
        chart_df["ghi"] = weatherhourly[ghi_col].reindex(hourly.index).values
    if weatherhourly is not None and temp_col:
        chart_df["temperature"] = weatherhourly[temp_col].reindex(hourly.index).values

    max_generation = max(float(chart_df["generationkwh"].max()), 1.0)
    tooltip_columns = [
        alt.Tooltip("time:T", title="시각"),
        alt.Tooltip("generationkwh:Q", title="발전량 [kWh]", format=",.2f"),
    ]
    series_frames = [
        pd.DataFrame(
            {
                "time": chart_df["time"],
                "series": "발전량",
                "plot_value": chart_df["generationkwh"],
                "actual_value": chart_df["generationkwh"],
                "unit": "kWh",
            }
        )
    ]
    if "ghi" in chart_df.columns:
        max_ghi = max(float(chart_df["ghi"].max()), 1.0)
        tooltip_columns.append(
            alt.Tooltip("ghi:Q", title="전일사량 [W/m²]", format=",.1f")
        )
        series_frames.append(
            pd.DataFrame(
                {
                    "time": chart_df["time"],
                    "series": "전일사량",
                    "plot_value": chart_df["ghi"].fillna(0.0) / max_ghi * max_generation,
                    "actual_value": chart_df["ghi"],
                    "unit": "W/m²",
                }
            )
        )
    if "temperature" in chart_df.columns:
        temp = chart_df["temperature"].fillna(0.0)
        temp_range = max(float(temp.max() - temp.min()), 1.0)
        tooltip_columns.append(
            alt.Tooltip("temperature:Q", title="온도 [°C]", format=",.1f")
        )
        series_frames.append(
            pd.DataFrame(
                {
                    "time": chart_df["time"],
                    "series": "온도",
                    "plot_value": (temp - float(temp.min())) / temp_range * max_generation,
                    "actual_value": chart_df["temperature"],
                    "unit": "°C",
                }
            )
        )
    long_df = pd.concat(series_frames, ignore_index=True)

    base = alt.Chart(long_df).encode(
        x=alt.X("time:T", title="날짜"),
        y=alt.Y("plot_value:Q", title="발전량 [kWh]", scale=alt.Scale(zero=True)),
        color=alt.Color(
            "series:N",
            title="범례",
            scale=alt.Scale(
                domain=["발전량", "전일사량", "온도"],
                range=["#1F77B4", "#FFB000", "#D62728"],
            ),
        ),
    )
    generation_area = (
        alt.Chart(long_df[long_df["series"] == "발전량"])
        .mark_area(opacity=0.28, color="#1F77B4")
        .encode(
            x=alt.X("time:T", title="날짜"),
            y=alt.Y("plot_value:Q", title="발전량 [kWh]", scale=alt.Scale(zero=True)),
        )
    )
    lines = base.mark_line(strokeWidth=2.5).encode(
        strokeDash=alt.StrokeDash(
            "series:N",
            legend=None,
            scale=alt.Scale(
                domain=["발전량", "전일사량", "온도"],
                range=[[], [5, 4], [2, 4]],
            ),
        )
    )
    hover_points = (
        alt.Chart(chart_df)
        .mark_point(opacity=0, size=260)
        .encode(x="time:T", y=alt.Y("generationkwh:Q"), tooltip=tooltip_columns)
    )
    chart = (
        alt.layer(generation_area, lines, hover_points)
        .properties(height=480)
        .interactive()
    )
    st.altair_chart(chart, use_container_width=True)


@st.dialog("손실 설정")
def render_loss_dialog(key_prefix):
    loss_labels = [
        ("soiling", "Soiling"),
        ("mismatch", "Mismatch"),
        ("wiring", "Wiring"),
        ("connections", "Connections"),
        ("lid", "LID"),
        ("nameplate_rating", "Nameplate"),
        ("age", "Age"),
        ("availability", "Availability loss"),
    ]
    for row_start in range(0, len(loss_labels), 2):
        row_cols = st.columns(2)
        for col, (key, label) in zip(row_cols, loss_labels[row_start:row_start + 2]):
            cur = st.session_state.form_losssettings.get(
                key, {"enabled": True, "value": float(LOSSPARAMS.get(key, 0.0))}
            )
            with col:
                enabled = st.checkbox(
                    label,
                    value=bool(cur.get("enabled", True)),
                    key=f"{key_prefix}_loss_dialog_en_{key}",
                    help=LOSS_HELP_TEXTS.get(key),
                )
                value = st.number_input(
                    "%",
                    value=float(cur.get("value", 0.0)),
                    disabled=not enabled,
                    key=f"{key_prefix}_loss_dialog_val_{key}",
                    label_visibility="collapsed",
                )
            st.session_state.form_losssettings[key] = {
                "enabled": enabled,
                "value": value,
            }

    if st.button("완료", type="primary", use_container_width=True):
        st.session_state.show_loss_dialog = False
        st.rerun()


def render_generator_input():
    gen = st.session_state.form_gen
    key_prefix = f"form_{st.session_state.get('form_version', 0)}"

    st.subheader("발전기 정보 입력")
    mode_text = (
        f"{st.session_state.form_edit_idx + 1}번 발전기 수정 중"
        if st.session_state.form_edit_idx is not None
        else "새 발전기 등록"
    )
    st.caption(mode_text)

    col1, col2, col3 = st.columns(3)

    with col1:
        st.markdown("##### 기본 정보")
        gen.name = st.text_input("이름", value=gen.name, key=f"{key_prefix}_name")

        st.markdown("##### 모듈 설정")
        gen.modulepdcstcw = st.number_input(
            "모듈 STC [W]",
            value=float(gen.modulepdcstcw),
            min_value=0.0,
            step=10.0,
            key=f"{key_prefix}_modulepdcstcw",
        )
        gen.modulecount = st.number_input(
            "모듈 개수",
            value=int(gen.modulecount),
            min_value=1,
            step=1,
            key=f"{key_prefix}_modulecount",
        )
        gen.gammapctperc = st.number_input(
            "온도계수 Gamma [%/°C]",
            value=float(gen.gammapctperc),
            step=0.01,
            key=f"{key_prefix}_gamma",
            help=FIELD_HELP_TEXTS["gamma"],
        )
        face_options = ["Monofacial", "Bifacial"]
        face_idx = face_options.index(gen.facetype) if gen.facetype in face_options else 1
        gen.facetype = st.selectbox(
            "모듈 타입", face_options, index=face_idx, key=f"{key_prefix}_facetype"
        )
        if gen.facetype == "Bifacial":
            gen.bifacialityfactorpct = st.number_input(
                "양면 발전 계수 [%]",
                value=float(getattr(gen, "bifacialityfactorpct", 70.0)),
                min_value=0.0,
                max_value=100.0,
                step=1.0,
                key=f"{key_prefix}_bifaciality",
                help=FIELD_HELP_TEXTS["bifaciality"],
            )

    with col2:
        st.markdown("##### 인버터 설정")
        gen.invacratedwper = (
            st.number_input(
                "인버터 1대 용량 [kW]",
                value=float(gen.invacratedwper) / 1000.0,
                min_value=0.0,
                step=0.1,
                key=f"{key_prefix}_invkw",
            )
            * 1000.0
        )
        gen.invertercount = st.number_input(
            "인버터 개수",
            value=int(gen.invertercount),
            min_value=1,
            step=1,
            key=f"{key_prefix}_invertercount",
        )
        gen.etainvnom = (
            st.number_input(
                "인버터 효율 [%]",
                value=float(gen.etainvnom) * 100.0,
                min_value=0.0,
                max_value=100.0,
                step=0.1,
                key=f"{key_prefix}_eta",
            )
            / 100.0
        )

        st.markdown("##### 설치 환경")
        gen.surfaceazimuth = st.number_input(
            "방위각 [deg]",
            value=float(gen.surfaceazimuth),
            step=1.0,
            key=f"{key_prefix}_surfaceazimuth",
            help=FIELD_HELP_TEXTS["azimuth"],
        )
        gen.surfacetilt = st.number_input(
            "경사각 [deg]",
            value=float(gen.surfacetilt),
            min_value=0.0,
            step=0.1,
            key=f"{key_prefix}_surfacetilt",
        )
        mount_options = ["Open rack", "Close mount"]
        mount_idx = mount_options.index(gen.mounting) if gen.mounting in mount_options else 0
        gen.mounting = st.selectbox(
            "설치 형태", mount_options, index=mount_idx, key=f"{key_prefix}_mounting"
        )
        gen.albedo = st.number_input(
            "알베도",
            value=float(gen.albedo),
            min_value=0.0,
            max_value=1.0,
            step=0.01,
            key=f"{key_prefix}_albedo",
            help=FIELD_HELP_TEXTS["albedo"],
        )
        gen.plannedavailability = (
            st.number_input(
                "계획 가동률 [%]",
                value=float(gen.plannedavailability) * 100.0,
                min_value=0.0,
                max_value=100.0,
                step=0.1,
                key=f"{key_prefix}_plannedavailability",
            )
            / 100.0
        )

    with col3:
        st.markdown("##### 장애물 및 손실 설정")
        with st.expander("장애물", expanded=True):
            if not st.session_state.form_obstacles:
                st.caption("등록된 장애물이 없습니다.")

            if st.button(
                "+ 장애물 추가",
                key=f"{key_prefix}_add_obstacle",
                use_container_width=True,
            ):
                st.session_state.form_obstacles.append(
                    {
                        "enabled": True,
                        "centerazdeg": 0.0,
                        "distm": 0.0,
                        "heightm": 0.0,
                        "widthm": 0.0,
                    }
                )
                st.rerun()

            for idx, obs in enumerate(st.session_state.form_obstacles):
                st.divider()
                st.markdown(f"**장애물 {idx + 1}**")
                obs["enabled"] = True
                c1, c2 = st.columns(2)
                obs["centerazdeg"] = c1.number_input(
                    "방위각 [deg]",
                    value=float(obs.get("centerazdeg") or 0.0),
                    key=f"{key_prefix}_obs_az_{idx}",
                )
                obs["distm"] = c2.number_input(
                    "거리 [m]",
                    value=float(obs.get("distm") or 0.0),
                    key=f"{key_prefix}_obs_dist_{idx}",
                )
                c3, c4 = st.columns(2)
                obs["heightm"] = c3.number_input(
                    "높이 [m]",
                    value=float(obs.get("heightm") or 0.0),
                    key=f"{key_prefix}_obs_height_{idx}",
                )
                obs["widthm"] = c4.number_input(
                    "너비 [m]",
                    value=float(obs.get("widthm") or 0.0),
                    key=f"{key_prefix}_obs_width_{idx}",
                )
                if st.button("장애물 삭제", key=f"{key_prefix}_delete_obstacle_{idx}"):
                    st.session_state.form_obstacles.pop(idx)
                    st.rerun()

        if st.button("손실 설정", key=f"{key_prefix}_open_loss_dialog", use_container_width=True):
            st.session_state.show_loss_dialog = True
            st.rerun()

        if st.session_state.get("show_loss_dialog", False):
            render_loss_dialog(key_prefix)

    st.divider()
    c_submit, c_reset = st.columns([2, 1])
    submit_label = "선택 항목 수정 완료" if st.session_state.form_edit_idx is not None else "+ 목록에 추가"
    if c_submit.button(submit_label, type="primary", use_container_width=True):
        new_gen = build_generator_from_form()
        if not new_gen.name.strip():
            st.error("발전기 이름을 입력해 주세요.")
        elif st.session_state.form_edit_idx is None:
            st.session_state.generators.append(new_gen)
            st.success(f"{new_gen.name} 발전기를 계산 대상 목록에 추가했습니다.")
            reset_form()
            st.rerun()
        else:
            st.session_state.generators[st.session_state.form_edit_idx] = new_gen
            st.success(f"{new_gen.name} 발전기 정보를 수정했습니다.")
            reset_form()
            st.rerun()

    if c_reset.button("입력 초기화", use_container_width=True):
        reset_form()
        st.rerun()


def render_results(start_date, end_date, lat, lon):
    if st.session_state.results is None:
        st.info("계산 대상 목록에서 시뮬레이션을 실행하면 결과가 여기에 표시됩니다.")
        return

    if st.session_state.get("capped_message"):
        st.warning(
            "Open-Meteo 예보 데이터 제공 범위 때문에 "
            f"{st.session_state.calc_end_date.strftime('%Y-%m-%d')}까지 계산했습니다."
        )

    st.subheader("시뮬레이션 결과")
    result_keys = list(st.session_state.results.keys())
    if "Total" in result_keys:
        result_keys = ["Total"] + [key for key in result_keys if key != "Total"]
    tabs = st.tabs(result_keys)

    weatherhourly = st.session_state.get("weatherhourly")
    ghi_col = first_existing_col(
        weatherhourly, ["shortwave_radiation", "shortwaveradiation"]
    ) if weatherhourly is not None else None
    temp_col = first_existing_col(
        weatherhourly, ["temperature_2m", "temperature2m"]
    ) if weatherhourly is not None else None

    for tab, key in zip(tabs, result_keys):
        with tab:
            hourly = st.session_state.results[key]["hourly"]
            daily = st.session_state.results[key]["daily"]
            total_kwh = daily["dailygenerationkwh"].sum()
            hourly_generation = hourly["generationkwh"]
            peak_kwh = float(hourly_generation.max())
            peak_time = hourly_generation.idxmax()
            avg_daily_kwh = float(daily["dailygenerationkwh"].mean())

            kpi1, kpi2, kpi3 = st.columns(3)
            kpi1.metric("총 발전량", f"{total_kwh:,.2f} kWh")
            kpi2.metric(
                "최대 발전량",
                f"{peak_kwh:,.2f} kWh",
                peak_time.strftime("%Y-%m-%d %H:%M"),
            )
            kpi3.metric("일평균 발전량", f"{avg_daily_kwh:,.2f} kWh/day")

            render_hourly_result_chart(hourly, weatherhourly, ghi_col, temp_col)

            st.markdown("##### 날짜별 시간대 발전량 [kWh]")
            df = hourly["generationkwh"].copy()
            try:
                df.index = df.index.tz_localize(None)
            except Exception:
                pass
            table_df = df.reset_index()
            table_df.columns = ["time", "generationkwh"]
            table_df["날짜"] = table_df["time"].dt.strftime("%Y-%m-%d")
            table_df["시간"] = table_df["time"].dt.hour
            mat = table_df.pivot_table(
                index="날짜",
                columns="시간",
                values="generationkwh",
                aggfunc="sum",
            ).reindex(columns=range(24)).fillna(0.0)
            mat.columns = [f"{hour:02d}시" for hour in mat.columns]
            st.dataframe(
                mat.style.format("{:.2f}").background_gradient(cmap="YlOrRd", axis=None),
                use_container_width=True,
                height=min(420, 74 + 35 * len(mat)),
            )

            st.markdown("##### 발전기별 세부 요약")
            detail_df = build_generator_detail_summary(st.session_state.results)
            st.dataframe(
                detail_df,
                use_container_width=True,
                hide_index=True,
            )

    st.divider()
    st.subheader("데이터 다운로드")
    excel_output = generate_excel_export(
        st.session_state.results, start_date, end_date, lat, lon
    )
    csv_output = build_csv_export(st.session_state.results)
    col_excel, col_csv = st.columns(2)
    with col_excel:
        st.download_button(
            label="Excel 결과 파일 다운로드",
            data=excel_output.getvalue(),
            file_name=f"PV_Forecast_{start_date.strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )
    with col_csv:
        st.download_button(
            label="CSV 시간별 데이터 다운로드",
            data=csv_output,
            file_name=f"PV_Forecast_Hourly_{start_date.strftime('%Y%m%d')}.csv",
            mime="text/csv",
            use_container_width=True,
        )


if "generators" not in st.session_state:
    st.session_state.generators = []
if "results" not in st.session_state:
    st.session_state.results = None
if "weatherhourly" not in st.session_state:
    st.session_state.weatherhourly = None
if "capped_message" not in st.session_state:
    st.session_state.capped_message = False
if "calc_end_date" not in st.session_state:
    st.session_state.calc_end_date = None
if "form_gen" not in st.session_state:
    reset_form()
if not st.session_state.get("empty_obstacle_default_applied", False):
    form_obstacles = st.session_state.get("form_obstacles", [])
    if len(form_obstacles) == 1 and is_old_default_obstacle(form_obstacles[0]):
        st.session_state.form_obstacles = []
        refresh_form_widgets()
    st.session_state.empty_obstacle_default_applied = True


with st.sidebar:
    st.header("위치 설정")
    lat = st.number_input("위도", value=37.4317862, format="%.7f")
    lon = st.number_input("경도", value=126.6485109, format="%.7f")

    st.divider()
    st.header("기간 설정")
    today = datetime.date.today()
    start_date = st.date_input("시작일", today)
    end_date = st.date_input("종료일", today + datetime.timedelta(days=7))
    st.caption("선택한 기간의 00:00부터 23:00까지 시간별로 계산합니다.")


st.title("PV-Forecast")
st.caption("태양광 발전량 예측 및 발전기별 시뮬레이션")

tab_input, tab_list, tab_result = st.tabs(
    ["발전기 정보 입력", "계산 대상 목록", "시뮬레이션 결과"]
)

with tab_input:
    render_generator_input()

with tab_list:
    st.subheader("계산 대상 목록")
    if not st.session_state.generators:
        st.info("아직 등록된 발전기가 없습니다. '발전기 정보 입력' 탭에서 발전기를 추가해 주세요.")
    else:
        summary = pd.DataFrame(generator_summary_rows())
        st.dataframe(summary, use_container_width=True, hide_index=True)

        st.divider()
        c_select, c_edit, c_delete = st.columns([1.2, 1, 1])
        with c_select:
            label_col, input_col = st.columns([0.9, 1.1], vertical_alignment="center")
            label_col.markdown("선택 번호")
            selected_no = input_col.number_input(
                "선택 번호",
                min_value=1,
                max_value=len(st.session_state.generators),
                value=1,
                step=1,
                label_visibility="collapsed",
            )
        selected_idx = int(selected_no) - 1

        if c_edit.button("선택 수정", use_container_width=True):
            load_form_from_generator(
                st.session_state.generators[selected_idx],
                edit_idx=selected_idx,
            )
            st.success(
                f"{selected_no}번 발전기를 수정 모드로 전환했습니다. "
                "'발전기 정보 입력' 탭을 확인해 주세요."
            )
            st.rerun()

        if c_delete.button("선택 삭제", type="secondary", use_container_width=True):
            deleted_name = st.session_state.generators[selected_idx].name
            st.session_state.generators.pop(selected_idx)
            if st.session_state.form_edit_idx == selected_idx:
                reset_form()
            st.success(f"{deleted_name} 발전기를 삭제했습니다.")
            st.rerun()

        st.divider()
        if st.button("전체 시뮬레이션 계산 실행", type="primary", use_container_width=True):
            if start_date > end_date:
                st.error("종료일은 시작일 이후여야 합니다.")
            else:
                max_forecast_date = today + datetime.timedelta(days=15)
                capped = False
                calc_end_date = end_date
                if end_date > max_forecast_date:
                    calc_end_date = max_forecast_date
                    capped = True

                st.session_state.capped_message = capped
                st.session_state.calc_end_date = calc_end_date

                with st.status("태양광 발전량 시뮬레이션 진행 중...", expanded=True) as status:
                    try:
                        start_s = start_date.strftime("%Y-%m-%d")
                        end_s = calc_end_date.strftime("%Y-%m-%d")

                        status.write("Open-Meteo 기상 데이터를 수집하는 중...")
                        tz, altitude_m = resolvetimezoneandelevation(lat, lon)
                        weatherhourly, weatherdaily = getweatherdatamixed(
                            lat, lon, start_s, end_s, tz
                        )

                        results = {}
                        total_hourly = None

                        for idx, gen in enumerate(st.session_state.generators, start=1):
                            status.write(f"{idx}. [{gen.name}] 발전량 계산 중...")
                            obstaclesenabled = [
                                obs for obs in gen.obstacles if obs.get("enabled")
                            ]
                            lossparams = lossparamsforgenerator(gen)
                            hourly, daily = computepvforgenerator(
                                weatherhourly,
                                gen,
                                obstaclesenabled,
                                lossparams,
                                lat,
                                lon,
                                tz,
                                altitude_m,
                            )
                            results[gen.name] = {"hourly": hourly, "daily": daily}

                            cur = hourly[["acpowerw", "generationkwh"]].copy()
                            total_hourly = cur if total_hourly is None else total_hourly.add(cur, fill_value=0.0)

                        status.write("전체 발전기 합계를 계산하는 중...")
                        total_daily = total_hourly["generationkwh"].resample("D").sum().to_frame(
                            name="dailygenerationkwh"
                        )
                        results["Total"] = {"hourly": total_hourly, "daily": total_daily}

                        st.session_state.results = results
                        st.session_state.weatherhourly = weatherhourly
                        status.update(label="계산 완료", state="complete", expanded=False)
                        st.success("'시뮬레이션 결과' 탭에서 결과를 확인해 주세요.")
                    except Exception as exc:
                        status.update(label="오류 발생", state="error", expanded=True)
                        st.error(f"오류 발생: {exc}")

with tab_result:
    render_results(start_date, end_date, lat, lon)
